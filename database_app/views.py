from django.shortcuts import render
from django.http import HttpResponse
from .models import Brand, Client, CommercialOrganization, Country, EstablishmentForm, Order, Product, Receipt, \
    SalesPoint
from .forms import BrandForm, ProductForm, CountryForm, EstablishmentFormForm, ClientForm, CommercialOrganizationForm, \
    OrderForm, ReceiptForm, SalesPointForm
from django.contrib import messages
from django.db import IntegrityError
from django.db.models import Sum
from django.db.models import Count, Q
from openpyxl import Workbook


def index(request):
    return render(request, 'index.html')


# Брэнд
def brand_table(request):
    brandForm = BrandForm()
    brands = Brand.objects.all().order_by('id_brand')

    if request.method == "POST":
        if request.POST.get('change_type') == "edit":
            id_brand = request.POST.get('id')
            new_brand_name = request.POST.get('name')
            edit_brand(id_brand, new_brand_name)
            messages.success(request, 'Данные обновлены')
        elif request.POST.get('change_type') == "add":
            new_brand_name = request.POST.get('name')
            add_brand(new_brand_name)
            messages.success(request, 'Данные обновлены')
        else:
            try:
                id_brand = request.POST.get('id')
                delete_brand(id_brand)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, 'Данные не обновлены, удаляемая запись имеет зависимости')

    return render(request, 'brand_table.html',
                  {'brandForm': brandForm,
                   'brands': brands})


def edit_brand(id_brand, brand_name):
    Brand.objects.filter(id_brand=id_brand).update(brand_name=brand_name)


def delete_brand(id_brand):
    Brand.objects.filter(id_brand=id_brand).delete()


def add_brand(brand_name):
    brand = Brand(brand_name=brand_name)
    brand.save()


# Товар
def product_table(request):
    productForm = ProductForm()
    products = Product.objects.all().order_by('id_product')
    brands = Brand.objects.all().order_by('id_brand')
    commercial_organizations = CommercialOrganization.objects.all().order_by('id_commercial_organization')
    countries = Country.objects.all().order_by('id_country')

    if request.method == "POST":
        if request.POST.get('change_type') == "filter":
            filter_product_type = request.POST.get('type')
            filter_product_size = request.POST.get('size')
            filter_fk_id_commercial_organization = request.POST.get('commercial_organization')
            filter_fk_id_brand = request.POST.get('brand')
            filter_fk_id_country = request.POST.get('country')
            if filter_product_type != "":
                products = products.filter(product_type=filter_product_type)
            if filter_product_size != "":
                products = products.filter(product_size=filter_product_size)
            if filter_fk_id_commercial_organization != "":
                products = products.filter(fk_id_commercial_organization=filter_fk_id_commercial_organization)
            if filter_fk_id_brand != "":
                products = products.filter(fk_id_brand=filter_fk_id_brand)
            if filter_fk_id_country != "":
                products = products.filter(fk_id_country=filter_fk_id_country)
            messages.success(request, 'Данные отфильтрованы')
        elif request.POST.get('change_type') == "edit":
            id_product = request.POST.get('id')
            new_product_type = request.POST.get('type')
            new_product_size = request.POST.get('size')
            new_fk_id_commercial_organization = request.POST.get('commercial_organization')
            new_fk_id_brand = request.POST.get('brand')
            new_fk_id_country = request.POST.get('country')
            edit_product(id_product, new_product_type, new_product_size,
                         new_fk_id_commercial_organization,
                         new_fk_id_brand,
                         new_fk_id_country)
            messages.success(request, 'Данные обновлены')
        elif request.POST.get('change_type') == "add":
            new_product_type = request.POST.get('type')
            new_product_size = request.POST.get('size')
            new_fk_id_commercial_organization = request.POST.get('commercial_organization')
            new_fk_id_brand = request.POST.get('brand')
            new_fk_id_country = request.POST.get('country')
            add_product(new_product_type, new_product_size,
                        new_fk_id_commercial_organization,
                        new_fk_id_brand,
                        new_fk_id_country)
            messages.success(request, 'Данные обновлены')
        else:
            try:
                id_product = request.POST.get('id')
                delete_product(id_product)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, 'Данные не обновлены, удаляемая запись имеет зависимости')

    # Созадим словарь для хранения записей моделей с замененными внешними ключами на другие данные дочерней модели
    products_with_related_data = []
    for product in products:
        data = {
            'id_product': product.id_product,
            'product_type': product.product_type,
            'product_size': product.product_size,
            'commercial_organization': product.fk_id_commercial_organization.organization_name,
            'brand': product.fk_id_brand.brand_name,
            'country': product.fk_id_country.country_name
        }
        products_with_related_data.append(data)

    return render(request, 'product_table.html',
                  {'productForm': productForm,
                   'products': products_with_related_data,
                   'brands': brands,
                   'commercial_organizations': commercial_organizations,
                   'countries': countries})


def edit_product(id_product, product_type, product_size, fk_id_commercial_organization, fk_id_brand, fk_id_country):
    Product.objects.filter(id_product=id_product).update(
        product_type=product_type,
        product_size=product_size,
        fk_id_commercial_organization=CommercialOrganization.objects.get(
            id_commercial_organization=fk_id_commercial_organization),
        fk_id_brand=Brand.objects.get(
            id_brand=fk_id_brand),
        fk_id_country=Country.objects.get(
            id_country=fk_id_country))


def delete_product(id_product):
    Product.objects.filter(id_product=id_product).delete()


def add_product(product_type, product_size, fk_id_commercial_organization, fk_id_brand, fk_id_country):
    product = Product(product_type=product_type,
                      product_size=product_size,
                      fk_id_commercial_organization=CommercialOrganization.objects.get(
                          id_commercial_organization=fk_id_commercial_organization),
                      fk_id_brand=Brand.objects.get(
                          id_brand=fk_id_brand),
                      fk_id_country=Country.objects.get(
                          id_country=fk_id_country))
    product.save()


# Страна
def country_table(request):
    countryForm = CountryForm()
    countries = Country.objects.all().order_by('id_country')

    if request.method == "POST":
        if request.POST.get('change_type') == "edit":
            id_country = request.POST.get('id')
            new_country_name = request.POST.get('name')
            edit_country(id_country, new_country_name)
            messages.success(request, 'Данные обновлены')
        elif request.POST.get('change_type') == "add":
            new_country_name = request.POST.get('name')
            add_country(new_country_name)
            messages.success(request, 'Данные обновлены')
        else:
            try:
                id_country = request.POST.get('id')
                delete_country(id_country)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, f'Данные не обновлены, удаляемая запись имеет зависимости')

    return render(request, 'country_table.html',
                  {'countryForm': countryForm,
                   'countries': countries})


def edit_country(id_country, country_name):
    Country.objects.filter(id_country=id_country).update(country_name=country_name)


def delete_country(id_country):
    Country.objects.filter(id_country=id_country).delete()


def add_country(country_name):
    country = Country(country_name=country_name)
    country.save()


# Форма учреждения
def establishment_form_table(request):
    establishment_formForm = EstablishmentFormForm()
    establishment_forms = EstablishmentForm.objects.all().order_by('id_establishment_form')

    if request.method == "POST":
        if request.POST.get('change_type') == "edit":
            id_establishment_form = request.POST.get('id')
            new_form_name = request.POST.get('name')
            edit_establishment_form(id_establishment_form, new_form_name)
        elif request.POST.get('change_type') == "add":
            new_form_name = request.POST.get('name')
            add_establishment_form(new_form_name)
        else:
            try:
                id_establishment_form = request.POST.get('id')
                delete_establishment_form(id_establishment_form)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, f'Данные не обновлены, удаляемая запись имеет зависимости')

    return render(request, 'establishment_form_table.html',
                  {'establishment_formForm': establishment_formForm,
                   'establishment_forms': establishment_forms})


def edit_establishment_form(id_establishment_form, form_name):
    EstablishmentForm.objects.filter(id_establishment_form=id_establishment_form).update(form_name=form_name)


def delete_establishment_form(id_establishment_form):
    EstablishmentForm.objects.filter(id_establishment_form=id_establishment_form).delete()


def add_establishment_form(form_name):
    establishment_form = EstablishmentForm(form_name=form_name)
    establishment_form.save()


# Клиент
def client_table(request):
    clientForm = ClientForm()
    clients = Client.objects.all().order_by('id_client')

    if request.method == "POST":
        if request.POST.get('change_type') == "edit":
            id_client = request.POST.get('id')
            new_client_name = request.POST.get('name')
            new_client_surname = request.POST.get('surname')
            new_client_patronymic = request.POST.get('patronymic')
            new_client_birth = request.POST.get('birth')
            new_client_phone = request.POST.get('phone')
            edit_client(id_client, new_client_name, new_client_surname, new_client_patronymic, new_client_birth,
                        new_client_phone)
            messages.success(request, 'Данные обновлены')
        elif request.POST.get('change_type') == "add":
            new_client_name = request.POST.get('name')
            new_client_surname = request.POST.get('surname')
            new_client_patronymic = request.POST.get('patronymic')
            new_client_birth = request.POST.get('birth')
            new_client_phone = request.POST.get('phone')
            add_client(new_client_name, new_client_surname, new_client_patronymic, new_client_birth, new_client_phone)
            messages.success(request, 'Данные обновлены')
        else:
            try:
                id_client = request.POST.get('id')
                delete_client(id_client)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, f'Данные не обновлены, удаляемая запись имеет зависимости')

    return render(request, 'client_table.html',
                  {'clientForm': clientForm,
                   'clients': clients})


def edit_client(id_client, client_name, client_surname, client_patronymic, client_birth,
                client_phone):
    Client.objects.filter(id_client=id_client).update(client_name=client_name, client_surname=client_surname,
                                                      client_patronymic=client_patronymic, client_birth=client_birth,
                                                      client_phone=client_phone)


def delete_client(id_client):
    Client.objects.filter(id_client=id_client).delete()


def add_client(client_name, client_surname, client_patronymic, client_birth, client_phone):
    client = Client(client_name=client_name, client_surname=client_surname,
                    client_patronymic=client_patronymic, client_birth=client_birth,
                    client_phone=client_phone)
    client.save()


# Коммерческая организация
def commercial_organization_table(request):
    commercial_organizationForm = CommercialOrganizationForm()
    commercial_organizations = CommercialOrganization.objects.all().order_by('id_commercial_organization')
    establishment_forms = EstablishmentForm.objects.all().order_by('id_establishment_form')

    if request.method == "POST":
        if request.POST.get('change_type') == "edit":
            id_commercial_organization = request.POST.get('id')
            new_organization_name = request.POST.get('name')
            new_link = request.POST.get('link')
            new_tin = request.POST.get('tin')
            new_fk_id_establishment_form = request.POST.get('establishment_form')
            edit_commercial_organization(id_commercial_organization, new_organization_name, new_link, new_tin,
                                         new_fk_id_establishment_form)
            messages.success(request, 'Данные обновлены')
        elif request.POST.get('change_type') == "add":
            new_organization_name = request.POST.get('name')
            new_link = request.POST.get('link')
            new_tin = request.POST.get('tin')
            new_fk_id_establishment_form = request.POST.get('establishment_form')
            add_commercial_organization(new_organization_name, new_link, new_tin,
                                        new_fk_id_establishment_form)
            messages.success(request, 'Данные обновлены')
        else:
            try:
                id_commercial_organization = request.POST.get('id')
                delete_commercial_organization(id_commercial_organization)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, 'Данные не обновлены, удаляемая запись имеет зависимости')

    # Созадим словарь для хранения записей моделей с замененными внешними ключами на другие данные дочерней модели
    commercial_organizations_with_related_data = []
    for commercial_organization in commercial_organizations:
        data = {
            'id_commercial_organization': commercial_organization.id_commercial_organization,
            'organization_name': commercial_organization.organization_name,
            'link': commercial_organization.link,
            'tin': commercial_organization.tin,
            'establishment_form': commercial_organization.fk_id_establishment_form.form_name
        }
        commercial_organizations_with_related_data.append(data)

    return render(request, 'commercial_organization_table.html',
                  {'commercial_organizationForm': commercial_organizationForm,
                   'commercial_organizations': commercial_organizations_with_related_data,
                   'establishment_forms': establishment_forms})


def edit_commercial_organization(id_commercial_organization, organization_name, link, tin, fk_id_establishment_form):
    CommercialOrganization.objects.filter(id_commercial_organization=id_commercial_organization).update(
        organization_name=organization_name, link=link,
        tin=tin,
        fk_id_establishment_form=EstablishmentForm.objects.get(
            id_establishment_form=fk_id_establishment_form))


def delete_commercial_organization(id_commercial_organization):
    CommercialOrganization.objects.filter(id_commercial_organization=id_commercial_organization).delete()


def add_commercial_organization(organization_name, link, tin, fk_id_establishment_form):
    commercial_organization = CommercialOrganization(organization_name=organization_name, link=link,
                                                     tin=tin,
                                                     fk_id_establishment_form=EstablishmentForm.objects.get(
                                                         id_establishment_form=fk_id_establishment_form))
    commercial_organization.save()


# Заказ
def order_table(request):
    orderForm = OrderForm()
    orders = Order.objects.all().order_by('id_order')
    clients = Client.objects.all().order_by('id_client')
    sales_points = SalesPoint.objects.all().order_by('id_sales_point')

    if request.method == "POST":
        if request.POST.get('change_type') == "filter":
            filter_order_address = request.POST.get('address')
            filter_allow_order_status = True if (request.POST.get('allow_status') == 'on') else False
            filter_order_status = True if (request.POST.get('status') == 'on') else False
            filter_order_date = request.POST.get('date')
            filter_fk_id_client = request.POST.get('client')
            filter_fk_id_sales_point = request.POST.get('sales_point')
            if filter_order_address != "":
                orders = orders.filter(order_address=filter_order_address)
            if filter_order_date != "":
                orders = orders.filter(order_date=filter_order_date)
            if filter_allow_order_status:
                orders = orders.filter(order_status=filter_order_status)
            if filter_fk_id_client != "":
                orders = orders.filter(fk_id_client=filter_fk_id_client)
            if filter_fk_id_sales_point != "":
                orders = orders.filter(fk_id_sales_point=filter_fk_id_sales_point)
            messages.success(request, 'Данные отфильтрованы')
        elif request.POST.get('change_type') == "edit":
            id_order = request.POST.get('id')
            new_order_address = request.POST.get('address')
            new_order_status = True if (request.POST.get('status') == 'on') else False
            new_order_date = request.POST.get('date')
            new_fk_id_client = request.POST.get('client')
            new_fk_id_sales_point = request.POST.get('sales_point')
            edit_order(id_order, new_order_address, new_order_status, new_order_date, new_fk_id_client,
                       new_fk_id_sales_point)
            messages.success(request, 'Данные обновлены')
        elif request.POST.get('change_type') == "add":
            new_order_address = request.POST.get('address')
            new_order_status = True if (request.POST.get('status') == 'on') else False
            new_order_date = request.POST.get('date')
            new_fk_id_client = request.POST.get('client')
            new_fk_id_sales_point = request.POST.get('sales_point')
            add_order(new_order_address, new_order_status, new_order_date, new_fk_id_client, new_fk_id_sales_point)
            messages.success(request, 'Данные обновлены')
        else:
            try:
                id_order = request.POST.get('id')
                delete_order(id_order)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, 'Данные не обновлены, удаляемая запись имеет зависимости')

    # Созадим словарь для хранения записей моделей с замененными внешними ключами на другие данные дочерней модели
    orders_with_related_data = []
    for order in orders:
        data = {
            'id_order': order.id_order,
            'order_address': order.order_address,
            'order_status': bool(order.order_status),
            'order_date': order.order_date,
            'client': f"{order.fk_id_client.client_surname} {order.fk_id_client.client_name} {order.fk_id_client.client_patronymic}",
            'sales_point': order.fk_id_sales_point.sales_point_address
        }
        orders_with_related_data.append(data)

    return render(request, 'order_table.html',
                  {'orderForm': orderForm,
                   'orders': orders_with_related_data,
                   'clients': clients,
                   'sales_points': sales_points})


def edit_order(id_order, order_address, order_status, order_date, fk_id_client, fk_id_sales_point):
    Order.objects.filter(id_order=id_order).update(
        order_address=order_address,
        order_status=order_status,
        order_date=order_date,
        fk_id_client=Client.objects.get(id_client=fk_id_client),
        fk_id_sales_point=SalesPoint.objects.get(id_sales_point=fk_id_sales_point))


def delete_order(id_order):
    Order.objects.filter(id_order=id_order).delete()


def add_order(order_address, order_status, order_date, fk_id_client, fk_id_sales_point):
    order = Order(order_address=order_address,
                  order_status=order_status,
                  order_date=order_date,
                  fk_id_client=Client.objects.get(id_client=fk_id_client),
                  fk_id_sales_point=SalesPoint.objects.get(id_sales_point=fk_id_sales_point))
    order.save()


# Чек
def receipt_table(request):
    receiptForm = ReceiptForm()
    receipts = Receipt.objects.all().order_by('id_receipt')
    orders = Order.objects.all().order_by('id_order')
    products = Product.objects.all().order_by('id_product')
    if request.method == "POST":
        if request.POST.get('change_type') == "filter":
            filter_receipt_product_amount = request.POST.get('amount')
            print(filter_receipt_product_amount)
            filter_receipt_product_price = request.POST.get('price')
            print(filter_receipt_product_price)
            filter_fk_id_order = request.POST.get('order')
            print(filter_fk_id_order)
            filter_fk_id_product = request.POST.get('product')
            print(filter_fk_id_product)
            if filter_receipt_product_amount != "":
                receipts = receipts.filter(receipt_product_amount=filter_receipt_product_amount)
            if filter_receipt_product_price != "":
                receipts = receipts.filter(receipt_product_price=filter_receipt_product_price)
            if filter_fk_id_product != "":
                receipts = receipts.filter(fk_id_product=filter_fk_id_product)
            if filter_fk_id_order != "":
                receipts = receipts.filter(fk_id_order=filter_fk_id_order)
            messages.success(request, 'Данные отфильтрованы')
        elif request.POST.get('change_type') == "edit":
            id_receipt = request.POST.get('id')
            new_receipt_product_amount = request.POST.get('amount')
            new_receipt_product_price = request.POST.get('price')
            new_fk_id_order = request.POST.get('order')
            new_fk_id_product = request.POST.get('product')
            edit_receipt(id_receipt, new_receipt_product_amount, new_receipt_product_price, new_fk_id_order,
                         new_fk_id_product)
            messages.success(request, 'Данные обновлены')
        elif request.POST.get('change_type') == "add":
            new_receipt_product_amount = request.POST.get('amount')
            new_receipt_product_price = request.POST.get('price')
            new_fk_id_order = request.POST.get('order')
            new_fk_id_product = request.POST.get('product')
            add_receipt(new_receipt_product_amount, new_receipt_product_price, new_fk_id_order, new_fk_id_product)
            messages.success(request, 'Данные обновлены')
        else:
            try:
                id_receipt = request.POST.get('id')
                delete_receipt(id_receipt)
                messages.success(request, 'Данные обновлены')
            except IntegrityError as e:
                messages.warning(request, 'Данные не обновлены, удаляемая запись имеет зависимости')

    return render(request, 'receipt_table.html',
                  {'receiptForm': receiptForm,
                   'receipts': receipts,
                   'orders': orders,
                   'products': products})


def edit_receipt(id_receipt, receipt_product_amount, receipt_product_price, fk_id_order, fk_id_product):
    Receipt.objects.filter(id_receipt=id_receipt).update(
        receipt_product_amount=receipt_product_amount,
        receipt_product_price=receipt_product_price,
        fk_id_order=Order.objects.get(id_order=fk_id_order),
        fk_id_product=Product.objects.get(id_product=fk_id_product))


def delete_receipt(id_receipt):
    Receipt.objects.filter(id_receipt=id_receipt).delete()


def add_receipt(receipt_product_amount, receipt_product_price, fk_id_order, fk_id_product):
    receipt = Receipt(receipt_product_amount=receipt_product_amount,
                      receipt_product_price=receipt_product_price,
                      fk_id_order=Order.objects.get(id_order=fk_id_order),
                      fk_id_product=Product.objects.get(id_product=fk_id_product))
    receipt.save()


def clients_total_receipt_price_report(request):
    # Получаем общую сумму заказов для каждого клиента
    clients_total_receipt_price = Client.objects.annotate(
        total_receipts_price=Sum('order__receipt__receipt_product_price'))
    # Добавляем выборочные значения в список
    clients_total_receipt_price_list = []
    for client in clients_total_receipt_price:
        clients_data = {
            "name": client.client_name,
            "surname": client.client_surname,
            "patronymic": client.client_patronymic,
            "total_price": client.total_receipts_price
            if client.total_receipts_price is not None else 0
        }
        clients_total_receipt_price_list.append(clients_data)

    if request.method == "POST":
        workbook = get_excel_workbook(clients_total_receipt_price_list)

        # Настройка HTTP-ответа для скачивания Excel-файла
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="clients_total_receipt_price_report.xlsx"'
        # Сохранение созданного Excel-файла в HTTP-ответе
        workbook.save(response)
        return response

    return render(request, 'clients_total_receipt_price_report.html', {'clients': clients_total_receipt_price_list})


def get_excel_workbook(objects_list):
    # Создание нового документа Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Данные'

    # Запись заголовков столбцов на первую строку
    headers = list(objects_list[0].keys())
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header

    # Запись данных из списка словарей в таблицу
    for row, data in enumerate(objects_list, start=2):
        for col, key in enumerate(data, start=1):
            sheet.cell(row=row, column=col).value = data[key]
    return workbook


def organizations_total_delivered_orders_report(request):
    # Получение количества доставленных заказов для каждой CommercialOrganization
    organizations_delivered_orders = CommercialOrganization.objects.annotate(
        delivered_orders_count=Count('salespoint__order', filter=Q(salespoint__order__order_status=True))
    ).values('organization_name', 'delivered_orders_count')

    organizations_delivered_list = []
    for organization in organizations_delivered_orders:
        organization_data = {
            "name": organization['organization_name'],
            "delivered_orders_amount": organization['delivered_orders_count']
        }
        organizations_delivered_list.append(organization_data)

    if request.method == "POST":
        workbook = get_excel_workbook(organizations_delivered_list)

        # Настройка HTTP-ответа для скачивания Excel-файла
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="organizations_delivered_list.xlsx"'
        # Сохранение созданного Excel-файла в HTTP-ответе
        workbook.save(response)
        return response

    return render(request, 'organizations_total_delivered_orders_report.html',
                  {'organizations': organizations_delivered_list})
